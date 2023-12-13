import openpyxl

class Funexcel():
    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file,path,sheetName):
        Worbook=openpyxl.load_workbook(path)
        sheet= Worbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(file,sheetName):
        Worbook= openpyxl.load_workbook(file)
        sheet=  Worbook[sheetName]
        return (sheet.max_column)

    def readData(file,path,sheetName,rownum,columnno):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file,path,sheetName,rownum,columnno,data):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Worbook.save(path)

    def get_test_data(self):
        # Abre el archivo Excel
        libro_excel = openpyxl.load_workbook("C:/Users/fpaez/OneDrive - DXC Production/Documents/Selenium/AutoSuccessfactor/TestData/Test_data.xlsx")
        hoja_excel = libro_excel.active

        # Lee datos desde Excel
        link_page = hoja_excel.cell(row=2, column=1).value
        user = hoja_excel.cell(row=2, column=2).value
        password = hoja_excel.cell(row=2, column=3).value
        id_empleado = hoja_excel.cell(row=2, column=4).value
        f_ingreso = hoja_excel.cell(row=2, column=5).value
        sociedad = hoja_excel.cell(row=2, column=6).value




        # Cierra el archivo Excel
        libro_excel.close()

        return link_page, user, password, id_empleado, f_ingreso, sociedad


